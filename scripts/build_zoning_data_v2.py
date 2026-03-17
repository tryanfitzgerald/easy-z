#!/usr/bin/env python3
"""
Build Zoning Data v2 — 5-Factor Scoring Model
Reads the research workbook and generates zoning_data.json for the heat map.
Can also be run standalone with hardcoded data for the initial 88 municipalities.
"""
import json, os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, 'data')

# Try to read from XLSX first, fall back to hardcoded
def load_from_xlsx():
    try:
        from openpyxl import load_workbook
        wb_path = os.path.join(DATA, 'easy_z_research_workbook.xlsx')
        if not os.path.exists(wb_path):
            return None
        wb = load_workbook(wb_path, data_only=True)
        ws = wb['Zoning Research']
        data = {}
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            state = ws.cell(row=row, column=2).value
            if not name or not state:
                continue
            key = f"{name}, {state}"

            zoning_score = ws.cell(row=row, column=7).value or 0
            site_score = ws.cell(row=row, column=17).value or 0
            approval_score = ws.cell(row=row, column=22).value or 0
            saturation_score = ws.cell(row=row, column=26).value or 0
            political_score = ws.cell(row=row, column=31).value or 0

            composite = ws.cell(row=row, column=32).value
            override = ws.cell(row=row, column=33).value
            override_reason = ws.cell(row=row, column=34).value or ''
            confidence = ws.cell(row=row, column=35).value or 'AI Estimate'

            dt_pathway = ws.cell(row=row, column=4).value or 'Unknown'
            cw_pathway = ws.cell(row=row, column=6).value or 'Unknown'
            stacking = ws.cell(row=row, column=8).value
            lot_min = ws.cell(row=row, column=9).value
            setback_front = ws.cell(row=row, column=10).value
            moratorium = ws.cell(row=row, column=28).value or 'Unknown'
            notes = ws.cell(row=row, column=30).value or ''

            final_score = override if override else (composite if composite else 0)
            if not final_score or final_score == 0:
                # Sum manually
                final_score = int(zoning_score) + int(site_score) + int(approval_score) + int(saturation_score) + int(political_score)

            try:
                final_score = int(final_score)
            except (TypeError, ValueError):
                final_score = 0

            if final_score == 0:
                continue  # Skip unresearched

            # Determine statuses from composite score
            def score_to_status(s):
                if s >= 85: return 'permitted'
                if s >= 67: return 'likely_permitted'
                if s >= 50: return 'conditional'
                if s >= 25: return 'restrictive'
                return 'prohibited'

            # For carwash vs drive-thru, use pathway info + composite
            cw_score = final_score  # Default same as composite
            dt_score = final_score

            # Adjust if we have pathway data
            if cw_pathway == 'Permitted by Right':
                cw_score = max(final_score, 85)
            elif cw_pathway == 'Prohibited':
                cw_score = min(final_score, 24)

            if dt_pathway == 'Permitted by Right':
                dt_score = max(final_score, 85)
            elif dt_pathway == 'Prohibited':
                dt_score = min(final_score, 24)

            data[key] = {
                'name': name,
                'state': state,
                'composite': final_score,
                'factors': {
                    'zoning': int(zoning_score) if zoning_score else 0,
                    'site_req': int(site_score) if site_score else 0,
                    'approval': int(approval_score) if approval_score else 0,
                    'saturation': int(saturation_score) if saturation_score else 0,
                    'political': int(political_score) if political_score else 0,
                },
                'carwash': {'status': score_to_status(cw_score), 'score': cw_score, 'pathway': str(cw_pathway)},
                'drivethru': {'status': score_to_status(dt_score), 'score': dt_score, 'pathway': str(dt_pathway)},
                'requirements': {
                    'stacking_cars': int(stacking) if stacking else None,
                    'lot_size_sqft': int(lot_min) if lot_min else None,
                    'front_setback_ft': int(setback_front) if setback_front else None,
                },
                'stacking_issue': (stacking and int(stacking) >= 10) if stacking else False,
                'moratorium': str(moratorium) == 'Yes',
                'confidence': str(confidence),
                'verified': str(confidence) in ('High', 'Medium'),
                'override_reason': str(override_reason) if override_reason else '',
                'notes': str(notes),
                'population': 0,  # Will be filled from master sheet
            }

        # Get population from master sheet
        ws_master = wb['Municipality Master']
        for row in range(2, ws_master.max_row + 1):
            name = ws_master.cell(row=row, column=1).value
            state = ws_master.cell(row=row, column=2).value
            pop = ws_master.cell(row=row, column=4).value
            if name and state:
                key = f"{name}, {state}"
                if key in data and pop:
                    data[key]['population'] = int(pop)

        return data
    except Exception as e:
        print(f'Could not read XLSX: {e}')
        return None

def main():
    data = load_from_xlsx()
    if data:
        print(f'Loaded {len(data)} municipalities from research workbook')
    else:
        print('XLSX not available, using existing zoning_data.json')
        with open(os.path.join(DATA, 'zoning_data.json')) as f:
            old = json.load(f)
        # Convert old format to new format
        data = {}
        for key, zd in old.items():
            old_cw = zd.get('carwash', {}).get('score', 50)
            old_dt = zd.get('drivethru', {}).get('score', 50)
            avg = (old_cw + old_dt) / 2
            factor_est = round(avg / 5)

            # Special case: Naperville
            if 'Naperville' in key:
                data[key] = {
                    'name': 'Naperville', 'state': 'IL',
                    'composite': 38,
                    'factors': {'zoning': 12, 'site_req': 8, 'approval': 5, 'saturation': 8, 'political': 5},
                    'carwash': {'status': 'conditional', 'score': 55, 'pathway': 'Conditional Use'},
                    'drivethru': {'status': 'restrictive', 'score': 38, 'pathway': 'Conditional Use'},
                    'requirements': {'stacking_cars': 10, 'lot_size_sqft': None, 'front_setback_ft': None},
                    'stacking_issue': True, 'moratorium': False,
                    'confidence': 'High', 'verified': True,
                    'override_reason': 'Ground truth: City blocked Dutch Bros on Napier Blvd due to 7 Brew at 1203 Iroquois',
                    'notes': 'Board actively hostile to new DT coffee concepts.',
                    'population': zd.get('population', 149000),
                }
                continue

            data[key] = {
                'name': zd.get('name', key.split(',')[0].strip()),
                'state': zd.get('state', key.split(',')[1].strip()),
                'composite': round(avg),
                'factors': {
                    'zoning': factor_est,
                    'site_req': factor_est,
                    'approval': factor_est,
                    'saturation': factor_est,
                    'political': factor_est,
                },
                'carwash': {'status': zd.get('carwash', {}).get('status', 'conditional'), 'score': old_cw, 'pathway': 'Unknown'},
                'drivethru': {'status': zd.get('drivethru', {}).get('status', 'conditional'), 'score': old_dt, 'pathway': 'Unknown'},
                'requirements': {'stacking_cars': None, 'lot_size_sqft': None, 'front_setback_ft': None},
                'stacking_issue': zd.get('stacking_issue', False),
                'moratorium': False,
                'confidence': 'AI Estimate',
                'verified': False,
                'override_reason': '',
                'notes': '',
                'population': zd.get('population', 0),
            }

    # Save
    out_path = os.path.join(DATA, 'zoning_data.json')
    with open(out_path, 'w') as f:
        json.dump(data, f, separators=(',', ':'))

    # Stats
    verified = sum(1 for v in data.values() if v['verified'])
    estimated = sum(1 for v in data.values() if not v['verified'])
    stacking = sum(1 for v in data.values() if v['stacking_issue'])
    print(f'Output: {out_path}')
    print(f'  Total: {len(data)} municipalities')
    print(f'  Verified: {verified}')
    print(f'  AI Estimated: {estimated}')
    print(f'  Stacking issues: {stacking}')

    # Show Naperville specifically
    if 'Naperville, IL' in data:
        nap = data['Naperville, IL']
        print(f'\n  Naperville, IL:')
        print(f'    Composite: {nap["composite"]}/100')
        print(f'    Drive-Thru: {nap["drivethru"]["score"]}/100 ({nap["drivethru"]["status"]})')
        print(f'    Carwash: {nap["carwash"]["score"]}/100 ({nap["carwash"]["status"]})')
        print(f'    Confidence: {nap["confidence"]}')

if __name__ == '__main__':
    main()
